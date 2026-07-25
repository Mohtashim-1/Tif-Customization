import json
from datetime import date
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import add_months, flt, formatdate, get_first_day, get_last_day, getdate
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

@frappe.whitelist()
def get_report_data(filters=None):
	filters = _parse_filters(filters)
	company = (filters.get("company") or "").strip()
	if not company:
		frappe.throw(_("Company is required."))

	default_from, default_to = _default_fiscal_dates(filters.get("reference_date"))
	from_date = getdate(filters.get("from_date") or default_from)
	to_date = getdate(filters.get("to_date") or default_to)

	if from_date > to_date:
		frappe.throw(_("From Date cannot be after To Date."))

	bank_accounts = _resolve_bank_accounts(filters, company)
	months = _month_range(from_date, to_date)
	month_keys = [m["key"] for m in months]

	banks = []
	for bank_account in bank_accounts:
		bank_account, gl_account, _company = _resolve_single_bank(bank_account, company)
		banks.append(
			_build_bank_series(
				bank_account, gl_account, company, from_date, to_date, months, month_keys
			)
		)

	donation_rows = []
	expense_rows = []
	for month in months:
		key = month["key"]
		donation_rows.append(
			{
				"month_label": month["label"],
				"banks": {b["bank_account"]: flt(b["donations"].get(key)) for b in banks},
			}
		)
		expense_rows.append(
			{
				"month_label": month["label"],
				"banks": {b["bank_account"]: flt(b["expenses"].get(key)) for b in banks},
			}
		)

	total_donation = sum(
		flt(b["donations"].get(key)) for b in banks for key in month_keys
	)
	total_expense = sum(
		flt(b["expenses"].get(key)) for b in banks for key in month_keys
	)
	donation_banks = {
		b["bank_account"]: sum(flt(b["donations"].get(key)) for key in month_keys) for b in banks
	}
	expense_banks = {
		b["bank_account"]: sum(flt(b["expenses"].get(key)) for key in month_keys) for b in banks
	}
	difference_banks = {
		bank_account: flt(donation_banks.get(bank_account)) - flt(expense_banks.get(bank_account))
		for bank_account in donation_banks
	}

	return {
		"title": _("Reconciliation with Bank Statements Summary"),
		"company": company,
		"from_date": str(from_date),
		"to_date": str(to_date),
		"from_date_label": formatdate(from_date, "dd MMM yyyy"),
		"to_date_label": formatdate(to_date, "dd MMM yyyy"),
		"months": months,
		"banks": [
			{
				"bank_account": b["bank_account"],
				"label": b["label"],
				"gl_account": b["gl_account"],
			}
			for b in banks
		],
		"initial_amount": {
			"date_label": formatdate(from_date, "dd MMM yyyy"),
			"banks": {b["bank_account"]: flt(b["initial"]) for b in banks},
		},
		"donations": donation_rows,
		"expenses": expense_rows,
		"ending_balance": {
			"date_label": formatdate(to_date, "dd MMM yyyy"),
			"banks": {b["bank_account"]: flt(b["ending"]) for b in banks},
		},
		"totals": {
			"total_donation": total_donation,
			"total_expense": total_expense,
			"total_difference": flt(total_donation) - flt(total_expense),
			"donation_banks": donation_banks,
			"expense_banks": expense_banks,
			"difference_banks": difference_banks,
		},
	}


@frappe.whitelist()
def get_bank_account_options(company=None, txt=""):
	company = (company or frappe.defaults.get_user_default("Company") or "").strip()
	if not company:
		return []

	filters = {"company": company, "disabled": 0, "is_company_account": 1}
	kwargs = {
		"filters": filters,
		"fields": ["name", "bank", "account_name"],
		"order_by": "bank, account_name, name",
		"limit_page_length": 0,
	}
	if txt:
		kwargs["or_filters"] = [
			["name", "like", f"%{txt}%"],
			["bank", "like", f"%{txt}%"],
			["account_name", "like", f"%{txt}%"],
		]
	rows = frappe.get_all("Bank Account", **kwargs)
	return [
		{
			"value": row.name,
			"description": _bank_account_label(row.name, row),
		}
		for row in rows
	]


@frappe.whitelist()
def download_reconciliation_excel(filters=None):
	report = get_report_data(filters=filters)
	wb = Workbook()
	ws = wb.active
	ws.title = "Reconciliation"

	bold = Font(bold=True)
	center = Alignment(horizontal="center", vertical="center", wrap_text=True)
	header_fill = PatternFill("solid", fgColor="E8EEF7")
	band_fill = PatternFill("solid", fgColor="B8C9E8")

	banks = report.get("banks") or []
	last_col = 2 + len(banks)

	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
	ws["A1"] = report.get("title")
	ws["A1"].font = Font(bold=True, size=14)
	ws["A1"].alignment = center

	row = 3
	ws.cell(row=row, column=1, value="")
	ws.cell(row=row, column=2, value="")
	col = 3
	for bank in banks:
		cell = ws.cell(row=row, column=col, value=bank.get("label"))
		cell.font = bold
		cell.fill = header_fill
		cell.alignment = center
		col += 1
	row += 1

	def write_amount_row(label, date_label, bank_values):
		nonlocal row
		ws.cell(row=row, column=1, value=label)
		ws.cell(row=row, column=2, value=date_label or "")
		col = 3
		for bank in banks:
			ws.cell(row=row, column=col, value=flt((bank_values or {}).get(bank["bank_account"])))
			col += 1
		row += 1

	def write_summary_row(label, bank_values):
		nonlocal row
		ws.cell(row=row, column=1, value=label)
		ws.cell(row=row, column=2, value="")
		col = 3
		for bank in banks:
			cell = ws.cell(row=row, column=col, value=flt((bank_values or {}).get(bank["bank_account"])))
			cell.font = bold
			col += 1
		for c in range(1, 3):
			ws.cell(row=row, column=c).font = bold
		row += 1

	initial = report.get("initial_amount") or {}
	write_amount_row(_("Initial Amount"), initial.get("date_label"), initial.get("banks"))

	donations = report.get("donations") or []
	if donations:
		start = row
		end = row + len(donations) - 1
		ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
		cell = ws.cell(row=start, column=1, value=_("Add donation"))
		cell.font = bold
		cell.fill = band_fill
		cell.alignment = Alignment(vertical="center", wrap_text=True)
		for item in donations:
			write_amount_row("", item.get("month_label"), item.get("banks"))
	write_summary_row(_("Total Donation"), (report.get("totals") or {}).get("donation_banks"))

	expenses = report.get("expenses") or []
	if expenses:
		start = row
		end = row + len(expenses) - 1
		ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
		cell = ws.cell(row=start, column=1, value=_("Less monthly payments & expenses"))
		cell.font = bold
		cell.fill = band_fill
		cell.alignment = Alignment(vertical="center", wrap_text=True)
		for item in expenses:
			write_amount_row("", item.get("month_label"), item.get("banks"))
	write_summary_row(_("Total Expense"), (report.get("totals") or {}).get("expense_banks"))

	totals = report.get("totals") or {}
	write_summary_row(_("Difference (Donation - Expense)"), totals.get("difference_banks"))

	ending = report.get("ending_balance") or {}
	write_amount_row(
		_("Ending balance {0}").format(ending.get("date_label") or ""),
		"",
		ending.get("banks"),
	)
	for c in range(1, last_col + 1):
		ws.cell(row=row - 1, column=c).font = bold

	for c in range(1, last_col + 1):
		ws.column_dimensions[get_column_letter(c)].width = 18 if c > 2 else 26

	output = BytesIO()
	wb.save(output)
	frappe.response["filename"] = "reconciliation_with_bank_statements.xlsx"
	frappe.response["filecontent"] = output.getvalue()
	frappe.response["type"] = "binary"


def _build_bank_series(bank_account, gl_account, company, from_date, to_date, months, month_keys):

	# Full GL keeps opening/ending cash position correct.
	erp_full = _gl_monthly_totals(gl_account, company, from_date, to_date, month_keys)
	erp_initial = _gl_balance_before(gl_account, company, from_date)

	# Donation / expense rows exclude internal bank transfers so they are not
	# counted as income on the receiving bank and expense on the sending bank.
	erp = _exclude_internal_transfers_from_monthly(
		erp_full, gl_account, company, from_date, to_date, month_keys
	)

	donations = {}
	expenses = {}
	for m in months:
		key = m["key"]
		donations[key] = flt((erp.get("debits") or {}).get(key))
		expenses[key] = flt((erp.get("credits") or {}).get(key))

	total_in = sum(flt(erp_full.get("debits", {}).get(k)) for k in month_keys)
	total_out = sum(flt(erp_full.get("credits", {}).get(k)) for k in month_keys)
	erp_ending = flt(erp_initial) + total_in - total_out

	return {
		"bank_account": bank_account,
		"label": _bank_account_label(bank_account),
		"gl_account": gl_account,
		"initial": flt(erp_initial),
		"donations": donations,
		"expenses": expenses,
		"ending": erp_ending,
	}


def _parse_filters(filters):
	if isinstance(filters, str):
		try:
			return json.loads(filters) or {}
		except Exception:
			return {}
	return filters or {}


def _default_fiscal_dates(reference=None):
	ref = getdate(reference or frappe.utils.nowdate())
	fy_start_year = ref.year if ref.month >= 7 else ref.year - 1
	from_date = date(fy_start_year, 7, 1)
	to_date = date(fy_start_year + 1, 4, 30)
	if ref < to_date:
		to_date = min(to_date, ref)
	return from_date, to_date


def _resolve_bank_accounts(filters, company):
	if cint_all(filters.get("all_bank_accounts")):
		return frappe.get_all(
			"Bank Account",
			filters={"company": company, "disabled": 0, "is_company_account": 1},
			pluck="name",
			order_by="bank, account_name, name",
		)

	raw = filters.get("bank_accounts") or filters.get("bank_account") or []
	if isinstance(raw, str):
		raw = [x.strip() for x in raw.split(",") if x.strip()]
	elif not isinstance(raw, list):
		raw = [raw] if raw else []

	bank_accounts = []
	for name in raw:
		if isinstance(name, dict):
			name = name.get("value") or name.get("name") or ""
		name = (name or "").strip()
		if name and name not in bank_accounts:
			bank_accounts.append(name)

	if not bank_accounts:
		frappe.throw(_("Select at least one Bank Account, or enable All Bank Accounts."))

	return bank_accounts


def cint_all(value):
	if isinstance(value, bool):
		return value
	if isinstance(value, (int, float)):
		return int(value)
	return str(value).strip().lower() in ("1", "true", "yes", "on")


def _resolve_single_bank(bank_account, company):
	row = frappe.db.get_value(
		"Bank Account",
		bank_account,
		["name", "account", "company", "disabled"],
		as_dict=True,
	)
	if not row:
		frappe.throw(_("Bank Account {0} was not found.").format(bank_account))
	if row.disabled:
		frappe.throw(_("Bank Account {0} is disabled.").format(bank_account))
	if not row.account:
		frappe.throw(_("Ledger account is not linked on Bank Account {0}.").format(bank_account))
	if row.company and row.company != company:
		frappe.throw(_("Bank Account {0} does not belong to company {1}.").format(bank_account, company))

	account = frappe.db.get_value("Account", row.account, ["name", "company", "account_type", "is_group"], as_dict=True)
	if not account:
		frappe.throw(_("Invalid ledger account {0}.").format(row.account))
	if account.is_group:
		frappe.throw(_("Please select a Bank Account with a non-group ledger account."))
	if account.account_type != "Bank":
		frappe.throw(_("Linked account {0} is not a Bank account.").format(row.account))

	return bank_account, row.account, company


def _bank_account_label(bank_account, row=None):
	if not row:
		row = frappe.db.get_value(
			"Bank Account",
			bank_account,
			["bank", "account_name", "name"],
			as_dict=True,
		)
	if not row:
		return bank_account
	parts = [p for p in [row.get("bank"), row.get("account_name")] if p]
	return " — ".join(parts) if parts else row.get("name") or bank_account


def _month_range(from_date, to_date):
	months = []
	cursor = get_first_day(from_date)
	end = get_last_day(to_date)
	while cursor <= end:
		key = cursor.strftime("%Y-%m")
		label = cursor.strftime("%b %y")
		months.append({"key": key, "label": label})
		cursor = add_months(cursor, 1)
	return months


def _empty_month_map(month_keys):
	return {k: 0.0 for k in month_keys}


def _gl_balance_before(gl_account, company, from_date):
	return flt(
		frappe.db.sql(
			"""
			SELECT COALESCE(SUM(debit - credit), 0)
			FROM `tabGL Entry`
			WHERE account = %(account)s
			  AND company = %(company)s
			  AND posting_date < %(from_date)s
			  AND IFNULL(is_cancelled, 0) = 0
			""",
			{"account": gl_account, "company": company, "from_date": from_date},
		)[0][0]
	)


def _gl_monthly_totals(gl_account, company, from_date, to_date, month_keys):
	debits = _empty_month_map(month_keys)
	credits = _empty_month_map(month_keys)
	rows = frappe.db.sql(
		"""
		SELECT
			DATE_FORMAT(posting_date, '%%Y-%%m') AS month_key,
			COALESCE(SUM(debit), 0) AS total_debit,
			COALESCE(SUM(credit), 0) AS total_credit
		FROM `tabGL Entry`
		WHERE account = %(account)s
		  AND company = %(company)s
		  AND posting_date BETWEEN %(from_date)s AND %(to_date)s
		  AND IFNULL(is_cancelled, 0) = 0
		GROUP BY month_key
		""",
		{"account": gl_account, "company": company, "from_date": from_date, "to_date": to_date},
		as_dict=True,
	)
	for row in rows:
		key = row.get("month_key")
		if key in debits:
			debits[key] = flt(row.get("total_debit"))
			credits[key] = flt(row.get("total_credit"))
	return {"debits": debits, "credits": credits}


def _exclude_internal_transfers_from_monthly(erp_full, gl_account, company, from_date, to_date, month_keys):
	"""Subtract internal bank-to-bank movements from donation/expense activity."""
	transfers = _internal_transfer_monthly_totals(gl_account, company, from_date, to_date, month_keys)
	debits = _empty_month_map(month_keys)
	credits = _empty_month_map(month_keys)
	for key in month_keys:
		debits[key] = max(flt(erp_full["debits"].get(key)) - flt(transfers["debits"].get(key)), 0)
		credits[key] = max(flt(erp_full["credits"].get(key)) - flt(transfers["credits"].get(key)), 0)
	return {"debits": debits, "credits": credits}


def _internal_transfer_monthly_totals(gl_account, company, from_date, to_date, month_keys):
	"""Monthly debit/credit on this bank from Internal Transfer PE and bank-to-bank JE."""
	debits = _empty_month_map(month_keys)
	credits = _empty_month_map(month_keys)
	params = {
		"account": gl_account,
		"company": company,
		"from_date": from_date,
		"to_date": to_date,
	}

	pe_rows = frappe.db.sql(
		"""
		SELECT
			DATE_FORMAT(gle.posting_date, '%%Y-%%m') AS month_key,
			COALESCE(SUM(gle.debit), 0) AS total_debit,
			COALESCE(SUM(gle.credit), 0) AS total_credit
		FROM `tabGL Entry` gle
		INNER JOIN `tabPayment Entry` pe
			ON pe.name = gle.voucher_no AND pe.docstatus = 1
		WHERE gle.account = %(account)s
		  AND gle.company = %(company)s
		  AND gle.voucher_type = 'Payment Entry'
		  AND gle.posting_date BETWEEN %(from_date)s AND %(to_date)s
		  AND IFNULL(gle.is_cancelled, 0) = 0
		  AND pe.payment_type = 'Internal Transfer'
		GROUP BY month_key
		""",
		params,
		as_dict=True,
	)
	for row in pe_rows:
		key = row.get("month_key")
		if key in debits:
			debits[key] += flt(row.get("total_debit"))
			credits[key] += flt(row.get("total_credit"))

	je_rows = frappe.db.sql(
		"""
		SELECT
			DATE_FORMAT(gle.posting_date, '%%Y-%%m') AS month_key,
			COALESCE(SUM(gle.debit), 0) AS total_debit,
			COALESCE(SUM(gle.credit), 0) AS total_credit
		FROM `tabGL Entry` gle
		WHERE gle.account = %(account)s
		  AND gle.company = %(company)s
		  AND gle.voucher_type = 'Journal Entry'
		  AND gle.posting_date BETWEEN %(from_date)s AND %(to_date)s
		  AND IFNULL(gle.is_cancelled, 0) = 0
		  AND (gle.debit > 0 OR gle.credit > 0)
		  AND EXISTS (
			SELECT 1
			FROM `tabGL Entry` contra
			INNER JOIN `tabAccount` acc ON acc.name = contra.account
			WHERE contra.voucher_type = gle.voucher_type
			  AND contra.voucher_no = gle.voucher_no
			  AND contra.company = gle.company
			  AND IFNULL(contra.is_cancelled, 0) = 0
			  AND contra.account != gle.account
			  AND acc.account_type = 'Bank'
			  AND (
				(gle.debit > 0 AND contra.credit > 0)
				OR (gle.credit > 0 AND contra.debit > 0)
			  )
		  )
		GROUP BY month_key
		""",
		params,
		as_dict=True,
	)
	for row in je_rows:
		key = row.get("month_key")
		if key in debits:
			debits[key] += flt(row.get("total_debit"))
			credits[key] += flt(row.get("total_credit"))

	return {"debits": debits, "credits": credits}
